from openpyxl import Workbook
import os
import json

wb = Workbook()
ws = wb.active

keys = ['id', 'stars', 'title', 'author_profile_url', 'author_name', 'badges', 'review_date', 'review_text', 'comments_count', 'review_helpful_votes']

with open(os.sys.argv[1]) as f:
    col = 1
    row = 1
    for key in keys:
        cell = ws.cell(row=row, column=col)
        cell.value = key
        col += 1
    row = 2
    for line in f:
        data = json.loads(line)
        col = 1

        for key in keys:
            cell = ws.cell(row=row, column=col)
            if type(data[key]) in [list,tuple]:
                cell.value = ', '.join(data[key])
            else:
                cell.value = '{}'.format(data[key])
            col += 1

        row += 1

wb.save(os.sys.argv[2])